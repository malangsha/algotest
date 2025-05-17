import logging
import os
import time
import threading
import queue
from enum import Enum
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import LineChart, Reference
import openpyxl.workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

from models.events import EventType, FillEvent, PositionEvent

class UpdateType(Enum):
    """Type of update to perform on the Excel dashboard."""
    ORDERS = "orders"
    TRADES = "trades"
    POSITIONS = "positions"
    SUMMARY = "summary"
    ALL = "all"

class ExcelDashboardMonitor:
    """
    Maintains an Excel dashboard showing orders, trades, positions, and summary metrics in real-time.
    Runs in a background thread to avoid blocking main program execution.
    Subscribes to events from an EventManager for live updates.
    """

    def __init__(self, position_manager, config: Dict[str, Any], event_manager=None):
        """
        Initialize the Excel dashboard monitor.

        Args:
            position_manager: The position manager instance.
            config: Global configuration dictionary.
            event_manager: Optional EventManager instance to subscribe to events.
        """
        self.position_manager = position_manager
        self.logger = logging.getLogger("core.excel_dashboard")
        
        # Configuration for the Excel monitor
        app_config = config or {}
        self.excel_config = app_config.get("excel_monitor", {})
        
        self.enabled = self.excel_config.get("enabled", False)
        self.excel_path = self.excel_config.get("path", "output/trading_dashboard.xlsx")
        self.update_interval = self.excel_config.get("update_interval", 2.0)
        self.max_retries = self.excel_config.get("max_retries", 5)
        self.retry_delay = self.excel_config.get("retry_delay", 3.0)

        # Display row counts from config
        display_rows_config = self.excel_config.get("display_rows", {})
        self.max_orders_display = display_rows_config.get("orders", 10)
        self.max_trades_display = display_rows_config.get("trades", 15)
        self.max_positions_display = display_rows_config.get("positions", 10)
        self.max_equity_points = display_rows_config.get("equity_points", 20)

        # Default column for equity‐curve data (will be hidden by the sheet setup)
        self.equity_data_start_col_letter = 'K'        
        self.equity_data_start_col = column_index_from_string(self.equity_data_start_col_letter)

        # Map each summary metric to its cell for quick updates        
        self.summary_metric_cells = {
            "initial_capital":      "B2",
            "cash_balance":         "C2",
            "portfolio_equity":     "D2",
            "total_realized_pnl":   "E2",
            "total_unrealized_pnl": "F2",
        }

        # Name of the Excel sheet used for the dashboard
        self.dashboard_sheet_name = "Dashboard"

        # Ensure output directory exists
        if self.enabled:
            try:
                os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            except OSError as e:
                self.logger.error(f"Could not create directory for Excel file {self.excel_path}: {e}")
                self.enabled = False # Disable if directory cannot be created

        # Threading and queue for non-blocking updates
        self.update_queue = queue.Queue()
        self.monitor_thread = None
        self.running = False
        
        # Data caches
        self.open_orders_data: List[Dict[str, Any]] = []
        self.executed_trades_data: List[Dict[str, Any]] = []
        self.positions_pnl_data: List[Dict[str, Any]] = []
        self.equity_curve_data: List[Tuple[datetime, float]] = [] # (timestamp, equity_value)
        
        # Portfolio metrics
        self.initial_capital = app_config.get("portfolio", {}).get("initial_capital", 1000000.0)
        self.cash_balance = self.initial_capital
        self.total_realized_pnl = 0.0
        self.total_unrealized_pnl = 0.0
        
        self.event_manager = event_manager

        if self.enabled:
            self.logger.info(f"Excel Dashboard Monitor initialized. Output: {self.excel_path}")
            self._create_or_load_dashboard_file()
            if self.event_manager:
                self._register_event_handlers()
            self.start()
        else:
            self.logger.info("Excel Dashboard Monitor is disabled in configuration.")

    def _register_event_handlers(self):
        """Register handlers for relevant events from the EventManager."""
        if not self.event_manager:
            return
        
        # Assuming EventType.FILL, EventType.POSITION, EventType.ORDER_STATUS exist
        # ORDER_STATUS would be for new open orders or changes to them.
        # For now, add_order is called manually or by another component.
        # If your system has an OrderManager that emits OrderStatusEvent, subscribe to it.
        # self.event_manager.subscribe(EventType.ORDER_STATUS, self.handle_order_event)
        
        self.event_manager.subscribe(EventType.FILL, self.handle_fill_event)
        self.event_manager.subscribe(EventType.POSITION, self.handle_position_event)
        self.logger.info("Registered event handlers for FILL and POSITION events.")

    def handle_fill_event(self, event: FillEvent):
        """Handles FillEvents to update trades and subsequently cash/positions."""
        if not self.enabled:
            return
        
        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Received FillEvent: {event.__dict__ if hasattr(event, '__dict__') else event}")

        trade_data = {
            'order_id': getattr(event, 'order_id', 'N/A'),
            'symbol': event.symbol,
            'side': str(getattr(event, 'side', 'N/A')), # Assuming side might be an enum
            'quantity': event.quantity,
            'fill_price': event.price, # Assuming event.price is fill_price
            'fill_time': datetime.fromtimestamp(event.timestamp / 1000).strftime("%Y-%m-%d %H:%M:%S") if hasattr(event, 'timestamp') else datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'commission': getattr(event, 'commission', 0.0)
        }
        self.add_trade(trade_data) # This will also trigger position and summary updates

    def handle_position_event(self, event: PositionEvent):
        """Handles PositionEvents to update P&L data."""
        if not self.enabled:
            return

        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Received PositionEvent: {event.__dict__ if hasattr(event, '__dict__') else event}")
        
        # A PositionEvent might be for a single symbol or portfolio-wide.
        # The current self.update_positions() fetches all positions.
        # This is robust.
        self.update_positions_and_pnl()

    def start(self):
        """Start the Excel dashboard monitor thread."""
        if not self.enabled:
            self.logger.info("Cannot start: Excel Dashboard Monitor is disabled.")
            return
            
        if self.monitor_thread and self.monitor_thread.is_alive():
            self.logger.warning("Excel monitor thread is already running.")
            return
            
        self.running = True
        self.monitor_thread = threading.Thread(
            target=self._monitor_thread_loop, 
            name="ExcelDashboardThread",
            daemon=True
        )
        self.monitor_thread.start()
        self.logger.info("Excel Dashboard Monitor thread started.")

    def stop(self):
        """Stop the Excel dashboard monitor thread."""
        self.running = False
        if self.monitor_thread and self.monitor_thread.is_alive():
            if self.logger.isEnabledFor(logging.DEBUG):
                self.logger.debug("Attempting to join Excel monitor thread...")
            self.update_queue.put(None) # Signal thread to exit
            self.monitor_thread.join(timeout=self.update_interval + 1.0)
            if self.monitor_thread.is_alive():
                self.logger.warning("Excel monitor thread did not terminate gracefully.")
            else:
                self.logger.info("Excel Dashboard Monitor thread stopped.")
        self.monitor_thread = None


    def _monitor_thread_loop(self):
        """Main loop for the Excel monitor thread."""
        retries = 0
        while self.running:
            try:
                update_task = self.update_queue.get(timeout=self.update_interval)
                if update_task is None: # Shutdown signal
                    break 
                
                update_type, data = update_task
                
                if self.logger.isEnabledFor(logging.DEBUG):
                    self.logger.debug(f"Processing update: {update_type}")

                if update_type == UpdateType.ORDERS:
                    self._add_item_to_cache(self.open_orders_data, data, self.max_orders_display)
                elif update_type == UpdateType.TRADES:
                    self._add_item_to_cache(self.executed_trades_data, data, self.max_trades_display)
                    # Update cash balance based on trade
                    qty = data.get('quantity', 0)
                    price = data.get('fill_price', 0)
                    side = data.get('side', '').upper()
                    commission = data.get('commission', 0.0)
                    
                    trade_value = qty * price
                    if side in ['SELL', 'SHORT']:
                        self.cash_balance += trade_value
                    else: # BUY, LONG
                        self.cash_balance -= trade_value
                    self.cash_balance -= commission # Commission always reduces cash

                elif update_type == UpdateType.POSITIONS:
                    self.positions_pnl_data = data # Expects a full list
                    # Recalculate total P&L from positions data
                    self.total_realized_pnl = sum(p.get('realized_PnL', 0) for p in self.positions_pnl_data)
                    self.total_unrealized_pnl = sum(p.get('unrealized_PnL', 0) for p in self.positions_pnl_data)

                # Always update Excel after processing a task from the queue
                self._update_excel_file()
                self.update_queue.task_done()
                retries = 0  # Reset retries on successful update
                
            except queue.Empty:
                # Timeout occurred, no item in queue, continue loop
                # This allows the thread to check self.running periodically
                if self.logger.isEnabledFor(logging.DEBUG):
                    self.logger.debug("Queue empty, continuing.")
                continue
            except PermissionError as pe:
                self.logger.warning(f"Excel file permission error (likely open): {pe}. Retrying later.")
                time.sleep(self.retry_delay * (1.5 ** retries))
                retries = min(retries + 1, self.max_retries)
            except Exception as e:
                retries += 1
                self.logger.error(f"Error in Excel monitor thread: {e}", exc_info=True)
                if retries > self.max_retries:
                    self.logger.critical(f"Excel monitor exceeded {self.max_retries} retries. Thread might be unstable. Stopping and restarting.")
                    # Simple restart logic: set running to false, then try to start again (if applicable from a manager)
                    # For a daemon thread, it might be better to log critical and let it die,
                    # relying on an external monitor to restart the whole app if needed.
                    # For now, we'll just log and continue, or break if it's too severe.
                    self.running = False # Stop the loop
                    self.logger.error("Excel monitor thread stopped due to excessive errors.")
                    break 
                else:
                    time.sleep(self.retry_delay * (1.5 ** retries)) # Exponential backoff
        self.logger.info("Excel monitor thread loop finished.")

    def _add_item_to_cache(self, cache: List, item: Dict, max_size: int):
        """Adds an item to the front of a cache list and trims to max_size."""
        cache.insert(0, item) # Add to the beginning for most recent first
        while len(cache) > max_size:
            cache.pop()

    def _create_or_load_dashboard_file(self):
        """Creates a new Excel dashboard file or loads an existing one for initial setup."""
        if not self.enabled:
            return

        # Define header styles
        self.header_font = Font(bold=True, color="FFFFFF") # White text
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue background
        self.section_title_font = Font(bold=True, size=14)
        self.centered_alignment = Alignment(horizontal='center', vertical='center')
        
        # Define PnL conditional formatting styles
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red

        try:
            if os.path.exists(self.excel_path):
                self.logger.info(f"Loading existing dashboard file: {self.excel_path}")
                # Try to open to ensure it's not corrupted
                wb = openpyxl.load_workbook(self.excel_path)
                if "Dashboard" not in wb.sheetnames:
                    self.logger.warning("Dashboard sheet not found. Recreating.")
                    self._initialize_dashboard_structure(wb)
                else:
                    # Optionally, clear old data from tables if desired on startup
                    pass
                wb.close()
            else:
                self.logger.info(f"Creating new dashboard file: {self.excel_path}")
                wb = openpyxl.Workbook()
                self._initialize_dashboard_structure(wb)
            
            # Perform an initial update to populate with any existing data
            self.update_all_sections_from_sources()

        except PermissionError:
            self.logger.error(f"Permission denied for Excel file {self.excel_path}. Ensure it's not open or locked.")
            self.enabled = False # Disable if file is inaccessible
        except Exception as e:
            self.logger.error(f"Error setting up dashboard file: {e}", exc_info=True)
            # Attempt to backup corrupted file and create new one
            if os.path.exists(self.excel_path):
                try:
                    backup_path = f"{self.excel_path}.bak.{int(time.time())}"
                    os.rename(self.excel_path, backup_path)
                    self.logger.info(f"Backed up potentially corrupt file to {backup_path}")
                    wb = openpyxl.Workbook()
                    self._initialize_dashboard_structure(wb)
                except Exception as backup_e:
                    self.logger.error(f"Failed to backup and recreate dashboard: {backup_e}")
                    self.enabled = False

    def _initialize_dashboard_structure(self, workbook: openpyxl.workbook):
        """Sets up the sheets, tables, and static formatting for the dashboard."""
       
        # Use the configured sheet name
        if self.dashboard_sheet_name in workbook.sheetnames:
            sheet = workbook[self.dashboard_sheet_name]
        else: 
            sheet = workbook.active
            sheet.title = self.dashboard_sheet_name

        # Clear all conditional formatting, tables, charts, merged cells if any        
        sheet.conditional_formatting._cf_rules = {}

        for table in list(sheet._tables): # Iterate over a copy
            del sheet._tables[table.name]
        for chart in list(sheet._charts):
            sheet._charts.remove(chart)
        
        # Unmerge all cells first
        merged_cells_ranges = list(sheet.merged_cells.ranges) # copy
        for merged_range in merged_cells_ranges:
            sheet.unmerge_cells(str(merged_range))

        # --- Dashboard Title ---
        sheet.merge_cells('A1:J1')
        title_cell = sheet['A1']
        title_cell.value = f"Trading Dashboard - Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.centered_alignment
        sheet.freeze_panes = 'A2' # Freeze title row

        current_row = 3 # Start content below title and a spacer row

        # --- Summary Metrics Section ---
        sheet.merge_cells(f'A{current_row}:J{current_row}')
        summary_title_cell = sheet.cell(row=current_row, column=1, value="PORTFOLIO SUMMARY")
        summary_title_cell.font = self.section_title_font
        summary_title_cell.alignment = self.centered_alignment
        summary_title_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light Gray
        current_row += 1

        summary_metrics = [
            ("Initial Capital:", self.initial_capital, 'B'),
            ("Current Cash Balance:", self.cash_balance, 'B'),
            ("Total Realized P&L:", self.total_realized_pnl, 'E'),
            ("Total Unrealized P&L:", self.total_unrealized_pnl, 'E'),
            ("Portfolio Equity:", self.cash_balance + self.total_unrealized_pnl, 'H') # Realized is already in cash
        ]
        
        metric_row_start = current_row
        for i, (label, value_placeholder, col_letter_val) in enumerate(summary_metrics):
            label_col = 'A' if i < 2 else ('D' if i < 4 else 'G') # Basic layout
            value_col = 'B' if i < 2 else ('E' if i < 4 else 'H')

            sheet.cell(row=current_row, column=openpyxl.utils.column_index_from_string(label_col), value=label).font = Font(bold=True)
            sheet.cell(row=current_row, column=openpyxl.utils.column_index_from_string(value_col), value=value_placeholder).number_format = '#,##0.00'
            if (i+1) % 2 == 0 and i < len(summary_metrics) -1 : # crude way to arrange in two rows
                 current_row +=1
            elif i == len(summary_metrics) -1:
                 current_row +=1


        # Store summary metric cell locations for easy updates
        self.summary_metric_cells = {
            "initial_capital": f'B{metric_row_start}',
            "cash_balance": f'B{metric_row_start + 1 if len(summary_metrics)>2 else metric_row_start}', # Adjust based on layout
            "total_realized_pnl": f'E{metric_row_start}',
            "total_unrealized_pnl": f'E{metric_row_start + 1 if len(summary_metrics)>2 else metric_row_start}',
            "portfolio_equity": f'H{metric_row_start}'
        }
        # Adjust if layout changes, for now, this is a simplified layout
        self.summary_metric_cells = {
            "initial_capital": 'B' + str(metric_row_start),
            "cash_balance": 'B' + str(metric_row_start + (1 if len(summary_metrics)>1 else 0)),
            "total_realized_pnl": 'E' + str(metric_row_start),
            "total_unrealized_pnl": 'E' + str(metric_row_start + (1 if len(summary_metrics)>1 else 0)),
            "portfolio_equity": 'H' + str(metric_row_start) # Assuming it's on the first row of metrics
        }
        current_row += 1 # Spacer

        # --- Equity Curve Chart Section ---
        eq_chart_start_row = current_row
        sheet.merge_cells(f'H{eq_chart_start_row}:J{eq_chart_start_row + self.max_equity_points + 1}') # Placeholder for chart
        sheet.cell(row=eq_chart_start_row, column=8, value="Equity Curve").font = Font(bold=True)
        # Data for equity curve will be in hidden columns or a separate sheet later if needed
        # For now, let's assume data in K, L (Time, Value)
        self.equity_data_start_col_letter = 'K'
        self.equity_data_start_col = openpyxl.utils.column_index_from_string(self.equity_data_start_col_letter)
        
        # Hide columns K and L which will store equity data
        sheet.column_dimensions[self.equity_data_start_col_letter].hidden = True
        sheet.column_dimensions[get_column_letter(self.equity_data_start_col + 1)].hidden = True


        # --- Table Sections ---
        # Open Orders
        current_row = self._create_table_section(sheet, current_row, "OPEN ORDERS",
                                     ['Order ID', 'Symbol', 'Side', 'Qty', 'Limit Price', 'Stop Price', 'Time Placed'],
                                     ['A', 'B', 'C', 'D', 'E', 'F', 'G'], # Column letters
                                     "OpenOrdersTable", self.max_orders_display)
        current_row += 2 # Spacer

        # Executed Trades
        current_row = self._create_table_section(sheet, current_row, "EXECUTED TRADES",
                                     ['Order ID', 'Symbol', 'Side', 'Qty', 'Fill Price', 'Commission', 'Fill Time'],
                                     ['A', 'B', 'C', 'D', 'E', 'F', 'G'],
                                     "ExecutedTradesTable", self.max_trades_display)
        current_row += 2 # Spacer

        # Positions & PnL
        current_row = self._create_table_section(sheet, current_row, "POSITIONS & PnL",
                                     ['Symbol', 'Net Qty', 'Avg Price', 'Unrealized PnL', 'Realized PnL'],
                                     ['A', 'B', 'C', 'D', 'E'],
                                     "PositionsTable", self.max_positions_display, pnl_cols=['D', 'E'])
        
        # Column Widths (approximate)
        col_widths = {'A': 18, 'B': 15, 'C': 10, 'D': 15, 'E': 15, 'F': 15, 'G': 20, 'H': 20, 'I':15, 'J':15 }
        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        try:
            workbook.save(self.excel_path)
            self.logger.info(f"Dashboard structure initialized and saved to {self.excel_path}")
        except PermissionError:
            self.logger.error(f"Permission denied saving initialized dashboard to {self.excel_path}. File might be open.")
            raise # Re-raise to be caught by caller

    def _create_table_section(self, sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int, title: str,
                              headers: List[str], col_letters: List[str], table_name: str,
                              max_data_rows: int, pnl_cols: Optional[List[str]] = None) -> int:
        """Helper to create a titled table section."""
        # Section Title
        end_col_letter = col_letters[-1]
        sheet.merge_cells(f'{col_letters[0]}{start_row}:{end_col_letter}{start_row}')
        title_cell = sheet.cell(row=start_row, column=openpyxl.utils.column_index_from_string(col_letters[0]), value=title)
        title_cell.font = self.section_title_font
        title_cell.alignment = self.centered_alignment
        title_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light Gray
        start_row += 1

        # Headers
        header_row = start_row
        for i, header_text in enumerate(headers):
            col_idx = openpyxl.utils.column_index_from_string(col_letters[i])
            cell = sheet.cell(row=header_row, column=col_idx, value=header_text)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        start_row += 1
        
        # Define table range (headers + max_data_rows)
        table_start_cell = f"{col_letters[0]}{header_row}"
        table_end_cell = f"{end_col_letter}{header_row + max_data_rows}"
        table_ref = f"{table_start_cell}:{table_end_cell}"
        
        excel_table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        excel_table.tableStyleInfo = style
        sheet.add_table(excel_table)

        # Conditional Formatting for PnL columns if specified
        if pnl_cols:
            for pnl_col_letter in pnl_cols:
                data_range = f"{pnl_col_letter}{start_row}:{pnl_col_letter}{start_row + max_data_rows -1}"
                # positive PnL → green
                sheet.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=self.green_fill)
                )
                
                # negative PnL → red
                sheet.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=self.red_fill)
                )

        return start_row + max_data_rows # Return next available row

    def _update_excel_file(self):
        """Writes the cached data to the Excel file."""
        if not self.enabled:
            return

        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Attempting to update Excel file: {self.excel_path}")

        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook["Dashboard"]

            # Update Dashboard Title Timestamp
            sheet['A1'].value = f"Trading Dashboard - Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Update Summary Metrics
            sheet[self.summary_metric_cells["initial_capital"]].value = self.initial_capital
            sheet[self.summary_metric_cells["cash_balance"]].value = self.cash_balance
            sheet[self.summary_metric_cells["total_realized_pnl"]].value = self.total_realized_pnl
            sheet[self.summary_metric_cells["total_unrealized_pnl"]].value = self.total_unrealized_pnl
            portfolio_equity = self.cash_balance + self.total_unrealized_pnl # Realized PNL is already part of cash balance
            sheet[self.summary_metric_cells["portfolio_equity"]].value = portfolio_equity
            
            # Update Equity Curve Data & Chart
            # Add current equity to history
            current_ts = datetime.now()
            self.equity_curve_data.append((current_ts, portfolio_equity))
            if len(self.equity_curve_data) > self.max_equity_points:
                self.equity_curve_data.pop(0) # Keep only the last N points

            # Clear previous equity data
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

            # summary_metric_cells["portfolio_equity"] is like "H4"
            col_letter, _row = coordinate_from_string(self.summary_metric_cells["portfolio_equity"])
            equity_data_start_row_abs = column_index_from_string(col_letter) + 3
            
            equity_data_col_idx = self.equity_data_start_col
            equity_data_start_sheet_row = sheet.tables['OpenOrdersTable'].ref.split(':')[0][1:] # Get row from table
            try:
                equity_data_start_sheet_row = int(equity_data_start_sheet_row) -1 -1 # Title, Header
                # This needs to be more robust, find the actual chart data area
                # For now, assume it's K3:L(3+max_equity_points)
                equity_data_start_sheet_row = 3 # Fixed for now, needs dynamic calc based on summary section
            except:
                equity_data_start_sheet_row = 3 # Fallback

            for i in range(self.max_equity_points):
                sheet.cell(row=equity_data_start_sheet_row + i, column=equity_data_col_idx).value = None # Time
                sheet.cell(row=equity_data_start_sheet_row + i, column=equity_data_col_idx + 1).value = None # Value
            
            for i, (ts, val) in enumerate(self.equity_curve_data):
                sheet.cell(row=equity_data_start_sheet_row + i, column=equity_data_col_idx).value = ts.strftime("%H:%M:%S")
                sheet.cell(row=equity_data_start_sheet_row + i, column=equity_data_col_idx + 1).value = val
                sheet.cell(row=equity_data_start_sheet_row + i, column=equity_data_col_idx + 1).number_format = '#,##0.00'

            self._update_or_create_equity_chart(sheet, equity_data_start_sheet_row)


            # Update Tables
            # Open Orders
            open_orders_table = sheet.tables["OpenOrdersTable"]
            self._update_table_data(sheet, open_orders_table, self.open_orders_data,
                                   ['order_id', 'symbol', 'side', 'quantity', 'limit_price', 'stop_price', 'time_placed'])
            # Executed Trades
            executed_trades_table = sheet.tables["ExecutedTradesTable"]
            self._update_table_data(sheet, executed_trades_table, self.executed_trades_data,
                                   ['order_id', 'symbol', 'side', 'quantity', 'fill_price', 'commission', 'fill_time'])
            # Positions & PnL
            positions_table = sheet.tables["PositionsTable"]
            self._update_table_data(sheet, positions_table, self.positions_pnl_data,
                                   ['symbol', 'net_quantity', 'avg_price', 'unrealized_PnL', 'realized_PnL'])

            workbook.save(self.excel_path)
            if self.logger.isEnabledFor(logging.DEBUG):
                self.logger.debug(f"Excel file updated successfully: {self.excel_path}")

        except PermissionError:
            self.logger.warning(f"Could not save Excel file (permission denied, likely open): {self.excel_path}")
        except KeyError as ke:  # Handles missing summary_metric_cells keys
            self.logger.error(
                f"Missing summary cell key: {ke}. Dashboard may be out of sync; re-initializing sheet.",
                exc_info=True
            )
            try:
                # Rebuild only the Dashboard sheet structure
                ws = workbook[self.dashboard_sheet_name]
                self._initialize_dashboard_structure(ws)
                workbook.save(self.excel_path)
            except Exception as e2:
                self.logger.error(f"Failed to re-initialize and save dashboard: {e2}", exc_info=True)
        finally:
            if 'workbook' in locals() and workbook:
                try:
                    workbook.close()
                except: # Might already be closed or in bad state
                    pass
    
    def _update_or_create_equity_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet, data_start_row: int):
        """Creates or updates the equity line chart."""
        chart_name = "EquityLineChart"
        
        # Remove existing chart if it exists to avoid duplicates
        existing_chart_index = -1
        for i, chart_obj in enumerate(sheet._charts):
            if chart_obj.title == chart_name or chart_obj.title == "Equity Curve": # Check for old title too
                existing_chart_index = i
                break
        if existing_chart_index != -1:
            del sheet._charts[existing_chart_index]

        if not self.equity_curve_data:
            return # No data to plot

        chart = LineChart()
        chart.title = chart_name
        chart.style = 13 # Choose a style
        chart.y_axis.title = 'Portfolio Equity'
        chart.x_axis.title = 'Time'
        chart.y_axis.crossAx = 500
        chart.x_axis = openpyxl.chart.axis.TextAxis(crossAx=100) # Treat X as category
        chart.x_axis.delete = False # Ensure X axis labels are shown
        chart.y_axis.delete = False # Ensure Y axis labels are shown
        chart.legend = None # No legend for a single series

        # Data for the chart (values)
        values_col_letter = get_column_letter(self.equity_data_start_col + 1)
        # values_ref_str = f"{values_col_letter}{data_start_row}:{values_col_letter}{data_start_row + len(self.equity_curve_data) - 1}"
        values_ref_str = f"{sheet.title}!{values_col_letter}{data_start_row}:{values_col_letter}{data_start_row}"
        values = Reference(sheet, range_string=values_ref_str)

        # Categories for the chart (timestamps)
        times_col_letter = get_column_letter(self.equity_data_start_col)
        # times_ref_str = f"{times_col_letter}{data_start_row}:{times_col_letter}{data_start_row + len(self.equity_curve_data) - 1}"
        times_ref_str = f"{sheet.title}!{times_col_letter}{data_start_row}:{times_col_letter}{data_start_row}"
        times = Reference(sheet, range_string=times_ref_str)

        series = openpyxl.chart.Series(values, title_from_data=False, title="Equity")
        chart.append(series)
        chart.set_categories(times)
        
        # Position the chart
        # Find where the "Equity Curve" title cell is. This is H{eq_chart_start_row}
        # Example: 'H5' if summary metrics take up to row 4
        # This needs to be calculated based on the actual layout.
        # For now, using a placeholder location.
        chart_anchor_cell = 'H' + str(sheet.tables['OpenOrdersTable'].ref.split(':')[0][1:]) # Anchor near first table
        try:
            # Attempt to find the summary section's end to place chart
            # This is a bit fragile and depends on the fixed layout of summary_metrics
            summary_end_row = openpyxl.utils.column_index_from_string(self.summary_metric_cells["portfolio_equity"][1:])
            chart_anchor_cell = f"H{summary_end_row + 2}"
        except:
             chart_anchor_cell = "H5" # Fallback anchor

        # A more robust way to find chart anchor:
        # Locate the "Equity Curve" cell created in _initialize_dashboard_structure
        # This cell was H{eq_chart_start_row}
        # Let's find that row dynamically.
        eq_chart_title_row = -1
        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row=r, column=8).value == "Equity Curve": # Column H is 8
                eq_chart_title_row = r
                break
        if eq_chart_title_row != -1:
             chart_anchor_cell = f"H{eq_chart_title_row + 1}"


        chart.width = 12 # Approx cm
        chart.height = 7  # Approx cm
        sheet.add_chart(chart, chart_anchor_cell)


    def _update_table_data(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                           table: openpyxl.worksheet.table.Table,
                           data_cache: List[Dict[str, Any]],
                           column_keys: List[str]):
        """Helper to clear and write data to an Excel table from a cache."""
        
        # Get table range (e.g., A3:G12)
        # The ref includes headers, so data starts one row below the start of ref
        table_start_cell, table_end_cell = table.ref.split(':')
        start_col_letter = ''.join(filter(str.isalpha, table_start_cell))
        start_row_abs = int(''.join(filter(str.isdigit, table_start_cell)))
        
        data_start_row = start_row_abs + 1 # Data rows start after header
        
        # Determine number of data rows in the table definition
        end_row_abs = int(''.join(filter(str.isdigit, table_end_cell)))
        max_table_data_rows = end_row_abs - start_row_abs

        # Clear existing data in the table's data rows
        for r_offset in range(max_table_data_rows):
            for c_offset in range(len(column_keys)):
                col_idx = openpyxl.utils.column_index_from_string(start_col_letter) + c_offset
                sheet.cell(row=data_start_row + r_offset, column=col_idx).value = None
                sheet.cell(row=data_start_row + r_offset, column=col_idx).number_format = 'General' # Reset format

        # Write new data from cache (most recent first, up to max_table_data_rows)
        for r_offset, record in enumerate(data_cache): # data_cache is already sorted newest first
            if r_offset >= max_table_data_rows:
                break
            current_sheet_row = data_start_row + r_offset
            for c_offset, key in enumerate(column_keys):
                col_idx = openpyxl.utils.column_index_from_string(start_col_letter) + c_offset
                value = record.get(key)
                cell = sheet.cell(row=current_sheet_row, column=col_idx, value=value)
                # Apply specific formatting if needed
                if isinstance(value, (int, float)):
                    if "price" in key.lower() or "pnl" in key.lower() or "commission" in key.lower():
                        cell.number_format = '#,##0.00'
                    elif "quantity" in key.lower() or "qty" in key.lower():
                        cell.number_format = '#,##0'
                if "time" in key.lower() and isinstance(value, str) and ":" in value: # Basic check for time string
                     cell.alignment = Alignment(horizontal='right')


    # --- Public methods to add/update data ---
    def add_open_order(self, order_data: Dict[str, Any]):
        """Adds a new open order to the dashboard."""
        if not self.enabled: return
        if 'time_placed' not in order_data:
            order_data['time_placed'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.update_queue.put((UpdateType.ORDERS, order_data))
        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Queued open order for update: {order_data.get('order_id')}")

    def add_trade(self, trade_data: Dict[str, Any]):
        """Adds a new executed trade to the dashboard."""
        if not self.enabled: return
        if 'fill_time' not in trade_data:
            trade_data['fill_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.update_queue.put((UpdateType.TRADES, trade_data))
        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Queued trade for update: {trade_data.get('order_id')}")
        
        # A trade affects positions and summary, so queue those updates too
        self.update_positions_and_pnl() # This will queue a POSITIONS update

    def update_positions_and_pnl(self):
        """Fetches current positions from PositionManager and queues an update."""
        if not self.enabled: return
        
        current_positions = []
        # Assuming position_manager.get_open_positions() returns a list of Position objects
        # And these objects have attributes like: instrument.symbol, quantity, entry_price, unrealized_pnl(), realized_pnl
        raw_positions = self.position_manager.get_open_positions()

        for pos in raw_positions:
            if pos.quantity == 0: # Don't show closed out positions here
                continue
            
            # Handle potential differences in Position object structure
            symbol = getattr(pos.instrument, 'symbol', str(pos.instrument))
            avg_price = getattr(pos, 'entry_price', getattr(pos, 'average_price', 0.0))
            unrealized_pnl_val = pos.unrealized_pnl() if callable(getattr(pos, 'unrealized_pnl', None)) else getattr(pos, 'unrealized_pnl', 0.0)
            realized_pnl_val = getattr(pos, 'realized_pnl', 0.0)

            current_positions.append({
                'symbol': symbol,
                'net_quantity': pos.quantity,
                'avg_price': avg_price,
                'unrealized_PnL': unrealized_pnl_val,
                'realized_PnL': realized_pnl_val, # This is per-position realized PnL
            })
        
        self.update_queue.put((UpdateType.POSITIONS, current_positions))
        if self.logger.isEnabledFor(logging.DEBUG):
            self.logger.debug(f"Queued positions update ({len(current_positions)} positions).")

    def update_all_sections_from_sources(self):
        """Manually trigger a refresh of all data sections from their sources."""
        if not self.enabled: return
        self.logger.info("Queueing full dashboard refresh.")
        # For orders, there's no source here, it's additive. So we just re-trigger write.
        # self.update_queue.put((UpdateType.ORDERS, {})) # This would clear if data is empty
        # For trades, also additive.
        # self.update_queue.put((UpdateType.TRADES, {}))
        self.update_positions_and_pnl() # This is the main one that pulls data.
        # The _update_excel_file will then update summary and chart based on these.

    def __del__(self):
        """Ensure thread is stopped on garbage collection."""
        if self.running:
            self.logger.info("ExcelDashboardMonitor being deleted, stopping thread...")
            self.stop()

# Example Usage (Illustrative - would be in your main application script)
if __name__ == '__main__':
    # --- Mockups for testing ---
    class MockInstrument:
        def __init__(self, symbol):
            self.symbol = symbol

    class MockPosition:
        def __init__(self, symbol, quantity, entry_price, realized_pnl, unrealized_pnl_val):
            self.instrument = MockInstrument(symbol)
            self.quantity = quantity
            self.entry_price = entry_price
            self.realized_pnl = realized_pnl
            self._unrealized_pnl_val = unrealized_pnl_val
        def unrealized_pnl(self):
            return self._unrealized_pnl_val

    class MockPositionManager:
        def __init__(self):
            self.positions = {}
        def add_mock_position(self, symbol, qty, entry, r_pnl, ur_pnl):
            self.positions[symbol] = MockPosition(symbol, qty, entry, r_pnl, ur_pnl)
        def get_open_positions(self):
            return [p for p in self.positions.values() if p.quantity != 0]
        def clear_positions(self):
            self.positions = {}

    class MockEventManager:
        def __init__(self):
            self.subscriptions = {}
        def subscribe(self, event_type, handler, component_name=None):
            if event_type not in self.subscriptions:
                self.subscriptions[event_type] = []
            self.subscriptions[event_type].append(handler)
            logging.info(f"MockEventManager: {component_name or 'Unknown'} subscribed to {event_type}")
        def publish(self, event):
            event_type = getattr(event, 'event_type', None) or type(event) # crude type detection
            if event_type in self.subscriptions:
                for handler in self.subscriptions[event_type]:
                    try:
                        handler(event)
                    except Exception as e:
                        logging.error(f"MockEventManager: Error in handler for {event_type}: {e}")
            else:
                 logging.debug(f"MockEventManager: No subscribers for event type {event_type}")


    # --- Setup Logging ---
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # --- Configuration (example) ---
    test_config = {
        "excel_monitor": {
            "enabled": True,
            "path": "output/test_trading_dashboard.xlsx",
            "update_interval": 1.0,
            "display_rows": {"orders": 5, "trades": 5, "positions": 5, "equity_points": 10}
        },
        "portfolio": {"initial_capital": 500000}
    }

    # --- Initialize ---
    mock_pos_mgr = MockPositionManager()
    mock_evt_mgr = MockEventManager()
    
    dashboard = ExcelDashboardMonitor(mock_pos_mgr, test_config, mock_evt_mgr)

    if not dashboard.enabled:
        logging.error("Dashboard could not be enabled. Exiting example.")
        exit()

    # --- Simulate some activity ---
    logging.info("Starting simulation...")

    # Simulate adding open orders
    dashboard.add_open_order({"order_id": "ORD001", "symbol": "RELIANCE", "side": "BUY", "quantity": 100, "limit_price": 2500.0, "stop_price": None})
    time.sleep(0.5)
    dashboard.add_open_order({"order_id": "ORD002", "symbol": "TCS", "side": "SELL", "quantity": 50, "limit_price": 3300.0, "stop_price": None})
    time.sleep(1.2) # Allow time for thread to process

    # Simulate a fill event (this would come from your broker/execution handler via EventManager)
    fill_event_1 = FillEvent(
        event_type=EventType.FILL, # Make sure your FillEvent has event_type or adjust publish
        symbol="RELIANCE", 
        order_id="ORD001", 
        side="BUY", 
        quantity=100, 
        price=2498.50, 
        timestamp=int(time.time() * 1000),
        commission=20.0
    )
    mock_evt_mgr.publish(fill_event_1) # This will call dashboard.handle_fill_event
    
    # Simulate position update after fill (PositionManager would do this and emit PositionEvent)
    mock_pos_mgr.add_mock_position("RELIANCE", 100, 2498.50, -20.0, 150.0) # (symbol, qty, entry, realized_pnl, unrealized_pnl)
    
    position_event_1 = PositionEvent(
        event_type=EventType.POSITION, # Make sure your PositionEvent has event_type
        symbol="RELIANCE", 
        quantity=100, 
        average_price=2498.50, 
        realized_pnl=-20.0, # PnL from this trade/fill
        unrealized_pnl=150.0, # Current unrealized PnL for the whole position
        timestamp=int(time.time() * 1000)
    )
    mock_evt_mgr.publish(position_event_1) # This will call dashboard.handle_position_event
    time.sleep(1.2)


    dashboard.add_open_order({"order_id": "ORD003", "symbol": "INFY", "side": "BUY", "quantity": 200, "limit_price": 1500.0, "stop_price": None})
    time.sleep(0.5)

    fill_event_2 = FillEvent(
        event_type=EventType.FILL,
        symbol="TCS", 
        order_id="ORD002", 
        side="SELL", 
        quantity=50, 
        price=3305.00, 
        timestamp=int(time.time() * 1000),
        commission=15.0
    )
    mock_evt_mgr.publish(fill_event_2)
    mock_pos_mgr.add_mock_position("TCS", -50, 3305.00, -15.0, -25.0) # Short position
    
    position_event_2 = PositionEvent(
        event_type=EventType.POSITION,
        symbol="TCS", 
        quantity=-50, 
        average_price=3305.00, 
        realized_pnl=-15.0,
        unrealized_pnl=-25.0, # Negative PnL
        timestamp=int(time.time() * 1000)
    )
    mock_evt_mgr.publish(position_event_2)
    time.sleep(1.2)

    # Simulate market data update causing PnL change for RELIANCE
    mock_pos_mgr.positions["RELIANCE"]._unrealized_pnl_val = 250.0 # Market moved favorably
    position_event_3 = PositionEvent(
        event_type=EventType.POSITION,
        symbol="RELIANCE", 
        quantity=100, 
        average_price=2498.50, 
        realized_pnl=-20.0, 
        unrealized_pnl=250.0, 
        timestamp=int(time.time() * 1000)
    )
    mock_evt_mgr.publish(position_event_3)
    time.sleep(1.2)

    logging.info("Simulation: Adding more orders to test scrolling/max display")
    for i in range(10):
        dashboard.add_open_order({
            "order_id": f"ORD{100+i}", "symbol": f"SYM{i}", "side": "BUY", 
            "quantity": 10+i, "limit_price": 100.0+i, "stop_price": None
        })
        time.sleep(0.1)
    time.sleep(1.2)


    logging.info("Simulation finished. Check the Excel file: output/test_trading_dashboard.xlsx")
    input("Press Enter to stop dashboard and exit...") # Keep running until Enter
    
    dashboard.stop()
    logging.info("Dashboard stopped.")
