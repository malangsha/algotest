import pytest
import os
import time
import openpyxl
from openpyxl.worksheet.table import Table
from datetime import datetime

# Assuming excel_dashboard_monitor is in the same directory or PYTHONPATH
from core.excel_dashboard_monitor import ExcelDashboardMonitor, UpdateType, EventType

# --- Mockups needed for testing ---
class MockInstrument:
    def __init__(self, symbol):
        self.symbol = symbol

class MockPosition:
    def __init__(self, symbol, quantity, entry_price, realized_pnl, unrealized_pnl_val):
        self.instrument = MockInstrument(symbol)
        self.quantity = quantity
        self.entry_price = entry_price
        self.realized_pnl = realized_pnl
        self._unrealized_pnl_val = unrealized_pnl_val
    def unrealized_pnl(self):
        return self._unrealized_pnl_val
    def __repr__(self):
        return f"<MockPosition {self.instrument.symbol} Qty:{self.quantity}>"


class MockPositionManager:
    def __init__(self):
        self.positions = {} # symbol -> MockPosition
        self.logger = logging.getLogger("mock.pos_mgr")

    def add_mock_position(self, symbol, qty, entry, r_pnl, ur_pnl):
        self.positions[symbol] = MockPosition(symbol, qty, entry, r_pnl, ur_pnl)
        self.logger.debug(f"Added/Updated mock position: {self.positions[symbol]}")


    def get_open_positions(self):
        self.logger.debug(f"Getting open positions. Current: {list(self.positions.values())}")
        return [p for p in self.positions.values() if p.quantity != 0]

    def apply_fill_to_mock(self, symbol, side, quantity, price, commission):
        # Simplified logic for mock:
        # If BUY, increase qty. If SELL, decrease. Update avg_price (not implemented here for simplicity).
        # Update realized PnL based on commission.
        current_pos = self.positions.get(symbol)
        trade_value = quantity * price
        
        if not current_pos:
            new_qty = quantity if side.upper() == "BUY" else -quantity
            self.add_mock_position(symbol, new_qty, price, -commission, 0) # Initial unrealized PnL is 0
        else:
            if side.upper() == "BUY":
                current_pos.quantity += quantity
            else: # SELL
                current_pos.quantity -= quantity
            current_pos.realized_pnl -= commission # Add commission to realized loss
            # Avg price and unrealized PnL update would be more complex.
        self.logger.debug(f"Applied fill to mock. New position for {symbol}: {self.positions.get(symbol)}")


class MockEventManager:
    def __init__(self):
        self.subscriptions = {}
        self.logger = logging.getLogger("mock.evt_mgr")
    def subscribe(self, event_type, handler, component_name=None):
        if event_type not in self.subscriptions:
            self.subscriptions[event_type] = []
        self.subscriptions[event_type].append(handler)
        self.logger.debug(f"Subscribed {component_name or 'Unnamed'} to {event_type}")
    def publish(self, event):
        event_type_to_check = getattr(event, 'event_type', type(event))
        self.logger.debug(f"Publishing event of type: {event_type_to_check}, data: {event.__dict__ if hasattr(event,'__dict__') else event}")
        if event_type_to_check in self.subscriptions:
            for handler in self.subscriptions[event_type_to_check]:
                handler(event)
        else:
            self.logger.debug(f"No subscribers for {event_type_to_check}")


# Placeholder events if not importable from models.events
try:
    from models.events import FillEvent, PositionEvent # Use actual events if available
except ImportError:
    class BaseEvent:
        def __init__(self, event_type, **kwargs):
            self.event_type = event_type
            self.__dict__.update(kwargs)

    class FillEvent(BaseEvent): # Placeholder
        def __init__(self, **kwargs):
            super().__init__(EventType.FILL, **kwargs)
            self.symbol = kwargs.get('symbol')
            self.quantity = kwargs.get('quantity')
            self.price = kwargs.get('price')
            self.side = kwargs.get('side')
            self.order_id = kwargs.get('order_id')
            self.timestamp = kwargs.get('timestamp', int(time.time() * 1000))
            self.commission = kwargs.get('commission', 0.0)


    class PositionEvent(BaseEvent): # Placeholder
        def __init__(self, **kwargs):
            super().__init__(EventType.POSITION, **kwargs)
            self.symbol = kwargs.get('symbol')
            self.quantity = kwargs.get('quantity')
            self.average_price = kwargs.get('average_price')
            self.realized_pnl = kwargs.get('realized_pnl')
            self.unrealized_pnl = kwargs.get('unrealized_pnl')
            self.timestamp = kwargs.get('timestamp', int(time.time() * 1000))


# --- Test Setup ---
TEST_XLSX_PATH = "output/pytest_dashboard.xlsx"
BASE_CONFIG = {
    "excel_monitor": {
        "enabled": True,
        "path": TEST_XLSX_PATH,
        "update_interval": 0.1, # Faster for tests
        "max_retries": 1,
        "retry_delay": 0.1,
        "display_rows": {"orders": 3, "trades": 3, "positions": 3, "equity_points": 5}
    },
    "portfolio": {"initial_capital": 100000}
}

# Configure logging for tests
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


@pytest.fixture
def mock_dependencies():
    pos_mgr = MockPositionManager()
    evt_mgr = MockEventManager()
    return pos_mgr, evt_mgr

@pytest.fixture
def dashboard_monitor(mock_dependencies):
    pos_mgr, evt_mgr = mock_dependencies
    # Clean up before test
    if os.path.exists(TEST_XLSX_PATH):
        os.remove(TEST_XLSX_PATH)
    
    monitor = ExcelDashboardMonitor(pos_mgr, BASE_CONFIG, evt_mgr)
    yield monitor # Provide the monitor to the test
    
    # Teardown: stop monitor and clean up file
    monitor.stop()
    time.sleep(0.2) # Give thread time to stop
    if os.path.exists(TEST_XLSX_PATH):
        try:
            os.remove(TEST_XLSX_PATH)
        except PermissionError: # File might be held by Excel briefly
            time.sleep(0.5)
            os.remove(TEST_XLSX_PATH)


# --- Test Cases ---
def test_dashboard_creation_and_structure(dashboard_monitor: ExcelDashboardMonitor):
    """Verify sheet existence, table headers, and basic formatting rules."""
    assert dashboard_monitor.enabled
    time.sleep(1) # Allow time for initial file creation and update

    assert os.path.exists(TEST_XLSX_PATH), "Excel file was not created"

    workbook = openpyxl.load_workbook(TEST_XLSX_PATH)
    assert "Dashboard" in workbook.sheetnames, "Dashboard sheet not found"
    sheet = workbook["Dashboard"]

    # Check main title
    assert "Trading Dashboard" in sheet['A1'].value

    # Check table existence and headers (example for OpenOrdersTable)
    assert "OpenOrdersTable" in [t.name for t in sheet._tables], "OpenOrdersTable not found"
    orders_table = sheet.tables["OpenOrdersTable"]
    
    # Table headers are in the first row of the table's range
    # e.g. if table.ref is "A5:G10", headers are in row 5
    table_start_cell, _ = orders_table.ref.split(':')
    header_row_idx = int(''.join(filter(str.isdigit, table_start_cell)))
    
    expected_order_headers = ['Order ID', 'Symbol', 'Side', 'Qty', 'Limit Price', 'Stop Price', 'Time Placed']
    for i, header in enumerate(expected_order_headers):
        col_letter = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string('A') + i)
        assert sheet[f"{col_letter}{header_row_idx}"].value == header

    # Check PnL conditional formatting (example for PositionsTable)
    # This is harder to check directly for applied rules, but we check if rules exist
    # For PositionsTable, PnL cols are D (Unrealized) and E (Realized)
    # Assuming PositionsTable headers start at some row, data starts 1 row below.
    # This requires knowing the exact layout, which is complex.
    # A simpler check: at least some conditional formatting rules exist.
    assert len(sheet.conditional_formatting.rules) > 0, "No conditional formatting rules found"
    
    # Check for summary metrics placeholders (example)
    # These cell locations are determined by _initialize_dashboard_structure
    # For this test, we'll just check if the portfolio equity cell has the initial capital
    initial_cap_cell_loc = dashboard_monitor.summary_metric_cells["initial_capital"]
    assert sheet[initial_cap_cell_loc].value == BASE_CONFIG["portfolio"]["initial_capital"]
    
    workbook.close()


def test_order_placement_and_fill_updates_dashboard(dashboard_monitor: ExcelDashboardMonitor, mock_dependencies):
    """Simulate an order placement + fill and assert correct row insertion."""
    pos_mgr, evt_mgr = mock_dependencies
    assert dashboard_monitor.enabled
    
    # 1. Add an open order
    order_data = {
        "order_id": "PYTEST001", "symbol": "TESTSYM", "side": "BUY", 
        "quantity": 10, "limit_price": 100.0, "stop_price": None,
        "time_placed": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    dashboard_monitor.add_open_order(order_data)
    time.sleep(0.5) # Allow queue processing

    # Verify order in Excel
    workbook = openpyxl.load_workbook(TEST_XLSX_PATH)
    sheet = workbook["Dashboard"]
    orders_table = sheet.tables["OpenOrdersTable"]
    data_start_row = int(orders_table.ref.split(':')[0][1:]) + 1 # Row after header
    assert sheet.cell(row=data_start_row, column=1).value == "PYTEST001" # Order ID in col A
    assert sheet.cell(row=data_start_row, column=2).value == "TESTSYM"   # Symbol in col B
    workbook.close()

    # 2. Simulate a fill for this order
    fill_data = FillEvent(
        symbol="TESTSYM", order_id="PYTEST001", side="BUY", 
        quantity=10, price=99.5, commission=1.0,
        timestamp=int(time.time()*1000)
    )
    # Publish event, ExcelDashboardMonitor is subscribed
    evt_mgr.publish(fill_data) 
    
    # The fill handler in dashboard_monitor will call add_trade, 
    # which then calls update_positions_and_pnl.
    # We need to ensure MockPositionManager reflects the new position for update_positions_and_pnl.
    pos_mgr.apply_fill_to_mock("TESTSYM", "BUY", 10, 99.5, 1.0) # Update mock PM
    
    # Simulate PositionManager emitting a PositionEvent after processing the fill
    # This is what triggers update_positions_and_pnl in the dashboard if subscribed.
    # Or, the add_trade method itself calls update_positions_and_pnl.
    # The current dashboard's add_trade calls update_positions_and_pnl(), so this should be covered.
    
    time.sleep(0.5) # Allow queue processing for trade and position update

    # Verify trade in Excel
    workbook = openpyxl.load_workbook(TEST_XLSX_PATH)
    sheet = workbook["Dashboard"]
    trades_table = sheet.tables["ExecutedTradesTable"]
    data_start_row_trades = int(trades_table.ref.split(':')[0][1:]) + 1
    assert sheet.cell(row=data_start_row_trades, column=1).value == "PYTEST001" # Order ID
    assert sheet.cell(row=data_start_row_trades, column=2).value == "TESTSYM"   # Symbol
    assert sheet.cell(row=data_start_row_trades, column=4).value == 10          # Quantity
    assert sheet.cell(row=data_start_row_trades, column=5).value == 99.5        # Fill Price

    # Verify position in Excel (after update_positions_and_pnl is triggered by add_trade)
    positions_table = sheet.tables["PositionsTable"]
    data_start_row_pos = int(positions_table.ref.split(':')[0][1:]) + 1
    assert sheet.cell(row=data_start_row_pos, column=1).value == "TESTSYM"      # Symbol
    assert sheet.cell(row=data_start_row_pos, column=2).value == 10             # Net Quantity
    assert sheet.cell(row=data_start_row_pos, column=3).value == 99.5           # Avg Price (from mock)
    assert sheet.cell(row=data_start_row_pos, column=5).value == -1.0           # Realized PnL (commission)
    
    # Verify cash balance update
    expected_cash = BASE_CONFIG["portfolio"]["initial_capital"] - (10 * 99.5) - 1.0 # capital - trade_val - commission
    cash_balance_cell_loc = dashboard_monitor.summary_metric_cells["cash_balance"]
    assert sheet[cash_balance_cell_loc].value == pytest.approx(expected_cash)

    workbook.close()
